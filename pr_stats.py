#!/usr/bin/env python3
"""PaddlePaddle 社区 PR 统计 — 纯云端版

通过 GitHub GraphQL API 逐仓库获取指定日期范围内已合并的 PR，
按仓库分组写入 Excel 模板，检测新贡献者，输出分类占比统计。

依赖: pip install requests openpyxl pyyaml
用法: python pr_stats.py --start 2026-03-12 --end 2026-03-19
"""

import argparse
import csv
import glob
import json
import os
import shutil
import sys
import time
from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple

import openpyxl
import requests

# ─── 路径 ────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


# ─── 配置加载 ────────────────────────────────────────────────

def load_config(config_path: Optional[str] = None) -> dict:
    """加载 repos.yaml 配置，如果不存在则用默认值"""
    path = config_path or os.path.join(SCRIPT_DIR, "repos.yaml")
    if os.path.exists(path):
        try:
            import yaml
            with open(path) as f:
                return yaml.safe_load(f)
        except ImportError:
            pass  # 没装 pyyaml，用默认值

    return {
        "org": "PaddlePaddle",
        "repos": [
            "Paddle", "docs", "PaddleScience", "PaddleMIX", "Paddle2ONNX",
            "PaddleOCR", "PaddleSpeech", "PaddleNLP", "FastDeploy",
            "GraphNet", "PaddleCustomDevice", "PaddleFormers", "PaddleX",
            "PaddleSOT", "PaConvert",
        ],
        "categories": ["社区", "硬件", "部门内", "部门外", "非硬件生态相关", "AI Coding"],
        "template": "社区开发者数据/commitlist_repo.xlsx",
        "output_dir": "社区开发者数据",
        "contributors": "contributors.csv",
    }


# ─── GitHub API ─────────────────────────────────────────────

def _get_proxies() -> dict:
    """从环境变量或 git config 获取代理配置"""
    for var in ("HTTPS_PROXY", "https_proxy", "HTTP_PROXY", "http_proxy"):
        val = os.environ.get(var, "")
        if val:
            return {"https": val, "http": val}
    import subprocess
    try:
        proxy = subprocess.check_output(
            ["git", "config", "--global", "https.proxy"],
            stderr=subprocess.DEVNULL,
        ).decode().strip()
        if proxy:
            return {"https": proxy, "http": proxy}
    except (subprocess.CalledProcessError, FileNotFoundError):
        pass
    return {}


def _get_token() -> str:
    """获取 GitHub token：环境变量 > ~/.gh_tokenrc > gh CLI"""
    for var in ("GITHUB_TOKEN", "GH_TOKEN"):
        val = os.environ.get(var, "")
        if val:
            return val
    token_file = os.path.expanduser("~/.gh_tokenrc")
    if os.path.exists(token_file):
        with open(token_file) as f:
            for line in f:
                if "github_oauth" in line and "=" in line:
                    return line.split("=", 1)[1].strip()
    print("错误: 未找到 GitHub token")
    print("  export GITHUB_TOKEN=ghp_xxx")
    print("  或: echo 'github_oauth = ghp_xxx' > ~/.gh_tokenrc")
    sys.exit(1)


def _graphql(token: str, query: str, retries: int = 3) -> dict:
    """带重试和限流处理的 GraphQL 请求"""
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    proxies = _get_proxies()
    for attempt in range(retries):
        resp = requests.post(
            "https://api.github.com/graphql",
            json={"query": query},
            headers=headers,
            timeout=30,
            proxies=proxies,
        )
        if resp.status_code == 200:
            data = resp.json()
            if "errors" in data:
                for err in data["errors"]:
                    if "rate limit" in err.get("message", "").lower():
                        reset = int(resp.headers.get("X-RateLimit-Reset", 0))
                        wait = max(reset - int(time.time()), 10)
                        print(f"  API 限流，等待 {wait}s ...")
                        time.sleep(wait)
                        continue
                print(f"GraphQL 错误: {data['errors']}")
                sys.exit(1)
            return data
        elif resp.status_code == 403:
            reset = int(resp.headers.get("X-RateLimit-Reset", 0))
            wait = max(reset - int(time.time()), 10)
            print(f"  API 限流 (403)，等待 {wait}s ...")
            time.sleep(wait)
        else:
            print(f"  HTTP {resp.status_code}，重试 {attempt+1}/{retries} ...")
            time.sleep(5)
    print(f"API 请求失败，已重试 {retries} 次")
    sys.exit(1)


def fetch_merged_prs(org: str, repos: List[str], start: str, end: str) -> Dict[str, List[dict]]:
    """逐仓库获取日期范围内已合并的 PR（使用 repository.pullRequests 端点，无 1000 条限制）"""
    token = _get_token()
    results = defaultdict(list)
    total = 0

    # 加载 ghost 别名
    ghost_aliases = _load_ghost_aliases()

    for repo in repos:
        cursor = None
        repo_count = 0
        while True:
            after_clause = f', after: "{cursor}"' if cursor else ""
            query = f'''{{
              repository(owner: "{org}", name: "{repo}") {{
                pullRequests(states: MERGED, first: 100, orderBy: {{field: CREATED_AT, direction: ASC}}{after_clause}) {{
                  totalCount
                  nodes {{
                    url
                    mergedAt
                    author {{ login }}
                  }}
                  pageInfo {{ hasNextPage endCursor }}
                }}
              }}
            }}'''

            data = _graphql(token, query)
            repo_data = data.get("data", {}).get("repository")
            if not repo_data:
                print(f"  {repo}: 仓库不存在或无权限，跳过")
                break

            pr_data = repo_data["pullRequests"]

            for node in pr_data["nodes"]:
                if not node:
                    continue
                merged_at = node.get("mergedAt", "")
                if not merged_at:
                    continue
                merge_date = merged_at[:10]
                # 按 mergedAt 过滤日期范围
                if merge_date < start or merge_date > end:
                    continue
                author = node.get("author")
                if author:
                    author_login = author["login"]
                else:
                    # ghost 账户：尝试从别名表恢复
                    url = node["url"]
                    author_login = ghost_aliases.get(url, "(ghost)")
                results[repo].append({"url": node["url"], "author": author_login})
                repo_count += 1

            if not pr_data["pageInfo"]["hasNextPage"]:
                break
            cursor = pr_data["pageInfo"]["endCursor"]

            # 优化：如果最早的 mergedAt 已经超过 end，后续不会有更多匹配
            last_merged = pr_data["nodes"][-1].get("mergedAt", "") if pr_data["nodes"] else ""
            if last_merged and last_merged[:10] > end:
                break

        if repo_count > 0:
            print(f"  {repo}: {repo_count} 条")
        total += repo_count

    print(f"合计 {total} 个已合并 PR（{len([r for r in results if results[r]])} 个仓库有数据）")
    return dict(results)


# ─── Ghost 别名 ─────────────────────────────────────────────

def _load_ghost_aliases() -> Dict[str, str]:
    """加载 ghost_aliases.csv（PR URL → 已知历史用户名）"""
    path = os.path.join(SCRIPT_DIR, "ghost_aliases.csv")
    if not os.path.exists(path):
        return {}
    with open(path, newline="") as f:
        return {row["url"]: row["author"] for row in csv.DictReader(f)}


# ─── 贡献者管理 ──────────────────────────────────────────────

def load_contributors(csv_path: str) -> Dict[str, str]:
    """加载 contributors.csv → {github_id: category}"""
    if not os.path.exists(csv_path):
        return {}
    with open(csv_path, newline="") as f:
        return {row["github_id"]: row["category"] for row in csv.DictReader(f)}


def detect_new_contributors(prs_by_repo: Dict[str, List[dict]], known: Dict[str, str]) -> List[dict]:
    """检测不在 contributors.csv 中的 PR 作者，返回 [{author, prs: [url, ...]}]"""
    author_prs: Dict[str, List[str]] = {}
    for prs in prs_by_repo.values():
        for pr in prs:
            if pr["author"] == "(ghost)":
                continue
            if pr["author"] not in known:
                author_prs.setdefault(pr["author"], []).append(pr["url"])
    return [{"author": a, "prs": urls} for a, urls in sorted(author_prs.items())]


def append_new_contributors(csv_path: str, new_ids: List[str]):
    """将新贡献者追加到 CSV，标记为 '待分类'"""
    with open(csv_path, "a", newline="") as f:
        w = csv.writer(f)
        for gid in new_ids:
            w.writerow([gid, "待分类"])


# ─── 统计 ────────────────────────────────────────────────────

def print_stats(prs_by_repo: Dict[str, List[dict]], repos: List[str],
                known: Dict[str, str], categories: List[str]):
    """输出分类统计和占比"""
    # 按仓库统计
    total_prs = 0
    print("\n各仓库 PR 数量:")
    for repo in repos:
        count = len(prs_by_repo.get(repo, []))
        if count > 0:
            print(f"  {repo}: {count}")
            total_prs += count
    print(f"  合计: {total_prs}")

    # 按分类统计
    cat_counts = defaultdict(int)  # type: Dict[str, int]
    for prs in prs_by_repo.values():
        for pr in prs:
            cat = known.get(pr["author"], "待分类")
            cat_counts[cat] += 1

    # 分类名 → 显示名映射
    display_map = {
        "部门内": "1-公司部门内",
        "部门外": "2-公司部门外",
        "硬件": "3-硬件公司",
        "社区": "4-个人贡献者",
        "非硬件生态相关": "5-非硬件生态相关",
        "AI Coding": "6-AI Coding",
        "待分类": "待分类",
    }

    print(f"\n分类占比（共 {total_prs} 条 PR）:")
    for cat in categories + ["待分类"]:
        n = cat_counts.get(cat, 0)
        if n > 0:
            pct = n / total_prs * 100 if total_prs else 0
            name = display_map.get(cat, cat)
            print(f"  {name}: {n} ({pct:.1f}%)")


# ─── sheet 名映射 ────────────────────────────────────────────

# 仓库名 → Excel sheet 名（年度文件约定）
REPO_TO_SHEET: Dict[str, str] = {
    "Paddle": "paddle",
    "docs": "docs",
    "PaddleScience": "science",
    "PaddleMIX": "mix",
    "Paddle2ONNX": "onnx",
    "PaddleOCR": "ocr",
    "PaddleSpeech": "Speech",
    "PaddleNLP": "NLP",
    "FastDeploy": "FastDeploy",
    "GraphNet": "GraphNet",
    "PaddleCustomDevice": "PaddleCustomDevice",
    "PaddleFormers": "PaddleFormers",
    "PaddleX": "PaddleX",
    "PaddleSOT": "PaddleSOT",
    "PaConvert": "PaConvert",
}


# ─── Excel 输出 ──────────────────────────────────────────────

def _get_annual_file(annual_file: str, repos: List[str]) -> openpyxl.Workbook:
    """加载或新建年度累积文件"""
    if os.path.exists(annual_file):
        return openpyxl.load_workbook(annual_file)
    # 新建：为每个仓库创建 sheet，写表头
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for repo in repos:
        sheet_name = REPO_TO_SHEET.get(repo, repo)
        ws = wb.create_sheet(sheet_name)
        ws.cell(1, 1, "")
        ws.cell(1, 2, "PR")
        ws.cell(1, 3, "贡献者")
        ws.cell(1, 4, "分类标注")
        ws.cell(1, 6, "社区贡献PR占比")
        ws.cell(1, 7, '=COUNTIF(D:D,"4-个人贡献者")/SUM(COUNTIF(D:D,"1-公司部门内"),COUNTIF(D:D,"2-公司部门外"),COUNTIF(D:D,"3-硬件公司"),COUNTIF(D:D,"4-个人贡献者"),COUNTIF(D:D,"5-非硬件生态相关"),COUNTIF(D:D,"6-AI Coding"))')
    return wb


def get_last_cutoff_from_annual(annual_file: str, primary_sheet: str = "paddle") -> Optional[str]:
    """从年度文件第一个有效 sheet 的最后截止标记检测上次截止日，返回次日"""
    if not os.path.exists(annual_file):
        return None
    wb = openpyxl.load_workbook(annual_file)
    # 优先用 primary_sheet，否则取第一个 sheet
    ws = wb[primary_sheet] if primary_sheet in wb.sheetnames else wb.worksheets[0]
    last_cutoff = None
    for r in range(ws.max_row, 0, -1):
        v = str(ws.cell(r, 1).value or "")
        if v.startswith("截止"):
            last_cutoff = v
            break
    if not last_cutoff:
        return None
    # 格式: 截止2026.04.02 或 截止2026年04月02日
    date_str = last_cutoff.replace("截止", "").strip()
    for fmt in ("%Y.%m.%d", "%Y年%m月%d日", "%Y.%-m.%-d"):
        try:
            d = datetime.strptime(date_str, fmt).date()
            return (d + timedelta(days=1)).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def append_to_annual(
    prs_by_repo: Dict[str, List[dict]],
    repos: List[str],
    annual_file: str,
    end_date: str,
    known: Dict[str, str],
):
    """将本周 PR 追加到年度累积文件，自动去重，末尾加截止标记行"""
    wb = _get_annual_file(annual_file, repos)
    cutoff_label = "截止" + datetime.strptime(end_date, "%Y-%m-%d").strftime("%Y.%m.%d")
    total_added = 0

    for repo in repos:
        prs = prs_by_repo.get(repo, [])
        if not prs:
            continue
        sheet_name = REPO_TO_SHEET.get(repo, repo)
        # 按需新建 sheet
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.cell(1, 2, "PR")
            ws.cell(1, 3, "贡献者")
            ws.cell(1, 4, "分类标注")
        else:
            ws = wb[sheet_name]

        # 收集已有 URL 用于去重
        existing_urls = set()
        for r in range(2, ws.max_row + 1):
            v = str(ws.cell(r, 2).value or "")
            if "github.com" in v:
                existing_urls.add(v)

        # 找下一个空行（跳过截止标记行）
        next_row = ws.max_row + 1

        added = 0
        for pr in prs:
            if pr["url"] in existing_urls:
                continue
            category = known.get(pr["author"], "待分类")
            ws.cell(next_row, 2, pr["url"])
            ws.cell(next_row, 3, pr["author"])
            ws.cell(next_row, 4, category)
            next_row += 1
            added += 1

        # 写截止标记行（即使本周没有新 PR 也写，标记已处理）
        ws.cell(next_row, 1, cutoff_label)
        total_added += added
        print(f"  {repo}({sheet_name}): +{added} 条")

    wb.save(annual_file)
    print(f"合计新增 {total_added} 条，截止标记: {cutoff_label}")
    return annual_file


def get_this_thursday() -> str:
    """获取本周四日期"""
    today = date.today()
    days_ahead = 3 - today.weekday()
    if days_ahead < 0:
        thu = today
    else:
        thu = today + timedelta(days=days_ahead)
    return thu.strftime("%Y-%m-%d")


def make_output_filename(end_date: str) -> str:
    """YYYY-MM-DD → commitlist_repo_YYYY_M_DD.xlsx"""
    d = datetime.strptime(end_date, "%Y-%m-%d")
    return f"commitlist_repo_{d.year}_{d.month}_{d.day}.xlsx"


# ─── 单调性守护 ──────────────────────────────────────────────

HISTORY_FILE = os.path.join(SCRIPT_DIR, "history.csv")


def _load_last_history() -> Optional[dict]:
    """读取 history.csv 最后一行"""
    if not os.path.exists(HISTORY_FILE):
        return None
    with open(HISTORY_FILE, newline="") as f:
        rows = list(csv.DictReader(f))
    return rows[-1] if rows else None


def _append_history(end_date: str, community_count: int, total_authors: int, ghost_count: int, new_this_week: int):
    """追加一行到 history.csv"""
    exists = os.path.exists(HISTORY_FILE)
    with open(HISTORY_FILE, "a", newline="") as f:
        w = csv.writer(f)
        if not exists:
            w.writerow(["date", "community_count", "total_authors", "ghost_count", "new_this_week"])
        w.writerow([end_date, community_count, total_authors, ghost_count, new_this_week])


def check_monotonicity(prs_by_repo: Dict[str, List[dict]], known: Dict[str, str], end_date: str, new_count: int) -> int:
    """检查社区开发者数是否单调不减，返回 ghost PR 数量"""
    # 统计 ghost PR 数
    ghost_count = sum(1 for prs in prs_by_repo.values() for pr in prs if pr["author"] == "(ghost)")

    # 统计所有唯一作者
    all_authors = {pr["author"] for prs in prs_by_repo.values() for pr in prs if pr["author"] != "(ghost)"}

    last = _load_last_history()
    if last:
        prev_count = int(last["community_count"])
        prev_date = last["date"]
        if new_count < prev_count:
            diff = prev_count - new_count
            print(f"\n⚠️  单调性告警: 社区开发者从 {prev_count}({prev_date}) 降至 {new_count}({end_date})，减少 {diff} 人")
            # 找出"丢失"的社区开发者
            current_community = {a for a in all_authors if known.get(a) == "社区"}
            # 无法直接比对历史（需要历史的 author 列表），打印当前数量供人工排查
            print(f"    当前社区作者数: {len(current_community)}")
            print(f"    Ghost PR 数: {ghost_count}")
            print(f"    建议运行 diagnose.py 做全量对比")

    _append_history(end_date, new_count, len(all_authors), ghost_count, new_count - int(last["community_count"]) if last else new_count)
    return ghost_count


# ─── 主入口 ──────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="PaddlePaddle 社区 PR 统计（年度累积版）")
    parser.add_argument("--start", help="起始日期 YYYY-MM-DD（留空则从年度文件最后截止日自动检测）")
    parser.add_argument("--end", help="截止日期 YYYY-MM-DD（留空则本周四）")
    parser.add_argument("--config", help="配置文件 repos.yaml（留空则用默认值）")
    parser.add_argument("--annual-file", help="年度累积文件路径（覆盖配置）")
    parser.add_argument("--contributors", help="贡献者列表路径（覆盖配置）")
    parser.add_argument("--dry-run", action="store_true", help="仅查询 API，不写文件")
    args = parser.parse_args()

    # 加载配置
    cfg = load_config(args.config)
    org = cfg["org"]
    repos = cfg["repos"]
    categories = cfg.get("categories", ["社区", "硬件", "部门内", "部门外", "非硬件生态相关", "AI Coding"])
    contributors_csv = args.contributors or os.path.join(SCRIPT_DIR, cfg["contributors"])

    # 1. 确定截止日期和年度文件路径
    end_date = args.end or get_this_thursday()
    year = datetime.strptime(end_date, "%Y-%m-%d").year
    annual_tpl = cfg.get("annual_file", "社区开发者pr统计（{year}）.xlsx")
    annual_file = args.annual_file or os.path.join(SCRIPT_DIR, annual_tpl.format(year=year))

    # 2. 确定起始日期：优先 --start，其次从年度文件最后截止标记检测
    if args.start:
        start_date = args.start
    else:
        start_date = get_last_cutoff_from_annual(annual_file)
        if not start_date:
            print("错误: 年度文件不存在或无截止标记，请用 --start 指定起始日期")
            sys.exit(1)

    print(f"日期范围: {start_date} → {end_date}")
    print(f"目标仓库: {len(repos)} 个 ({', '.join(repos[:3])}, ...)")
    print(f"年度文件: {annual_file}")

    # 3. 获取 PR 数据
    prs_by_repo = fetch_merged_prs(org, repos, start_date, end_date)

    # 4. 检测新贡献者
    known = load_contributors(contributors_csv)
    new_contribs = detect_new_contributors(prs_by_repo, known)
    if new_contribs:
        new_ids = [c["author"] for c in new_contribs]
        print(f"\nNEW_CONTRIBUTORS: {','.join(new_ids)}")
        print(f"发现 {len(new_contribs)} 个新贡献者:")
        for c in new_contribs:
            print(f"  - {c['author']}")
            for url in c["prs"]:
                print(f"      {url}")
        append_new_contributors(contributors_csv, new_ids)
        print(f"已追加到 {contributors_csv}（标记为「待分类」）")
        known = load_contributors(contributors_csv)
    else:
        print("\n无新贡献者")

    # 5. 输出统计
    print_stats(prs_by_repo, repos, known, categories)

    # 5.5 单调性检查 + 审计日志
    community_authors = set()
    for prs in prs_by_repo.values():
        for pr in prs:
            if known.get(pr["author"]) == "社区":
                community_authors.add(pr["author"])
    community_count = len(community_authors)
    check_monotonicity(prs_by_repo, known, end_date, community_count)

    if args.dry_run:
        print("\n[DRY RUN] 不写入文件")
        return

    # 6. 追加到年度文件
    print(f"\n写入年度文件:")
    append_to_annual(prs_by_repo, repos, annual_file, end_date, known)


if __name__ == "__main__":
    main()
